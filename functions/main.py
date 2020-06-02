import tkinter as tk
from tkinter import *
from openpyxl import load_workbook
import openpyxl
from variableCollector import *
from variableSet import *
from runCalc import *
from setExport import export_dialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


#Lectura de excel de materiales
materials = load_workbook(filename = 'resources/sheets/materials.xlsx')
sheet_obj = materials.active 
max_col = sheet_obj.max_row 
matList = [None]
for i in range(3, max_col + 1): 
    j=i-3
    matList.append(sheet_obj.cell(row = i, column = 3).value)

#Inicializacion GUI
root=tk.Tk()
root.title('acustica 2')
root.geometry('800x800')

#Componentes GUI

#Filtro
filtrado = StringVar()
filtrado.set('Tercios') 
selectFiltrado = OptionMenu(root, filtrado, 'Tercios','Octavas').grid(column=0,row=1)
Label(root, text="Filtrado").grid(column=0,row=1)

#Material
material = StringVar()
material.set('Acero')
Label(root, text="Material").grid(column=0,row=2) 
selectMaterial = OptionMenu(root, material, *matList).grid(column=0,row=2)

#Medidas
length = DoubleVar()
length.set(1)
Label(root, text="Largo").grid(column=0,row=3)
lengthInput = Entry(root, textvariable = length).grid(column=0,row=3)

height = DoubleVar()
height.set(1)
Label(root, text="Alto").grid(column=0,row=4)
heightInput = Entry(root, textvariable = height).grid(column=0,row=4)

thickness = DoubleVar()
thickness.set(1)
Label(root, text="Espesor").grid(column=0,row=5)
thicknessInput = Entry(root, textvariable = thickness).grid(column=0,row=5)

#Métodos
davy = IntVar()
Checkbutton(root, text="Davy", variable=davy).grid(column=0,row=6)

sharp = IntVar()
Checkbutton(root, text="Sharp", variable=sharp).grid(column=0,row=7)
cremer = IntVar()
Checkbutton(root, text="Cremer", variable=cremer).grid(column=0,row=8)

iso = IntVar()
Checkbutton(root, text="ISO", variable=iso).grid(column=0,row=9)

runFilter = tk.Button(root, text='run', fg='white', bg='#263D42', command= lambda:run_Calc(filtrado.get(), 
                                                                                                  length.get(),
                                                                                                  height.get(),
                                                                                                  thickness.get(),
                                                                                                  material.get(),
                                                                                                  davy.get(),
                                                                                                  sharp.get(),
                                                                                                  cremer.get(),
                                                                                                  iso.get(),
                                                                                                  chart,
                                                                                                  P
                                                                                                  ))
runFilter.grid(column=0,row=10)
runExport = tk.Button(root, text='Export', fg='white', bg='#263D42', command= lambda: export_dialog())
runExport.grid(column=0,row=11)

#Graph

fig = plt.Figure(figsize=(5,10), dpi=100)
P = fig.add_subplot()
chart = FigureCanvasTkAgg(fig,root)
chart.get_tk_widget().grid(column=0,row=1,columns0an=1,rowspan=11,sticky=NE)
chart.columnconfigure(1, weight=1)
P.grid()
    




root.mainloop()

