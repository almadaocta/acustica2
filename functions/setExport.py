from openpyxl import load_workbook
import openpyxl
from os.path import dirname, abspath
from saveExport import export_save
from openpyxl.chart import Reference, Series, LineChart

export={}
def set_export(calcData,Rd,Rs,Rc,Ri,ly):

    f=calcData['filterVector'] 
    material=calcData['material']
    thickness=calcData['thickness']
    height=calcData['height']
    lenght=calcData['length']
    Fc=calcData['Fc']

    export['file'] = load_workbook(filename = 'resources/sheets/exportTemplate.xlsx')
    sheets = export['file'].sheetnames
    wb = export['file'] 
    ws=wb[sheets[0]]
    ws.cell(row = 2, column = 1).value = material
    ws.cell(row = 2, column = 2).value = lenght
    ws.cell(row = 2, column = 3).value = height
    ws.cell(row = 2, column = 4).value = thickness
    ws.cell(row = 2, column = 5).value = Fc
    if ly==1:
        ws.cell(row = 2, column = 7).value = calcData['material2']
        ws.cell(row = 2, column = 8).value = calcData['thickness2']
    for i in range(0, len(calcData['filterVector'])):
         ws.cell(row = 4, column = i+2).value = f[i]
    for i in range(0, len(calcData['filterVector'])):
        ws.cell(row = 5, column = i+2).value = Rc[i]   
        ws.cell(row = 6, column = i+2).value = Rs[i]
        ws.cell(row = 7, column = i+2).value = Rd[i]
        ws.cell(row = 8, column = i+2).value = Ri[i]     

    #Graph

    chart = LineChart()
    chart.title = 'TL'
    maxCol=len(f)+2
    xAxis = Reference(ws,min_col=2, min_row=4, max_col=maxCol, max_row=4)
    # Cremer
    values = Reference(ws,min_col=1, min_row=5, max_col=maxCol, max_row=5)
    series = Series(values,title_from_data=True)
    chart.append(series)
    # Sharp
    values = Reference(ws,min_col=1, min_row=6, max_col=maxCol, max_row=6)
    series = Series(values,title_from_data=True)
    chart.append(series)
    # Davy
    values = Reference(ws,min_col=1, min_row=7, max_col=maxCol, max_row=7)
    series = Series(values,title_from_data=True)
    chart.append(series)
    # ISO
    values = Reference(ws,min_col=1, min_row=8, max_col=maxCol, max_row=8)
    series = Series(values,title_from_data=True)
    chart.append(series)

    chart.set_categories(xAxis)
    chart.height = 15 # default is 7.5
    chart.width = 25 # default is 15
    chart.x_axis.title = 'f[Hz]'
    chart.y_axis.title = 'Db'
    ws.add_chart(chart, 'A10') 


    #Hoja 2
    ws=wb[sheets[1]]
    #for i in range(26, 26+len(calcData['isoVector'])):
    #    ws.cell(row = i+1, column = 4).value = calcData['isoVector'][i-26]
    #    ws.cell(row = i+1, column = 16).value = calcData['isoVector'][i-26]

    for i in range(26, 26+len(calcData['ralt'])):
        ws.cell(row = i+1, column = 4).value = round(calcData['ralt'][i-26],1)
        ws.cell(row = i+1, column = 16).value = round(calcData['ralt'][i-26],0)

    for i in range(10, 10+len(calcData['rwCurve'])):
        ws.cell(row = 2, column = i).value = round(calcData['rwCurve'][i-10],1)
        ws.cell(row = 3, column = i).value = round(calcData['StcCurve'][i-10],0)
    

    ws.cell(row = 2, column = 2).value=calcData['Rw']
    ws.cell(row = 3, column = 2).value=calcData['C']
    ws.cell(row = 4, column = 2).value= calcData['Ctr']
    ws.cell(row = 5, column = 2).value=calcData['C503150']
    ws.cell(row = 6, column = 2).value=calcData['C505000']
    ws.cell(row = 7, column = 2).value=calcData['C1005000']
    ws.cell(row = 8, column = 2).value=calcData['CTR501350']
    ws.cell(row = 9, column = 2).value=calcData['CTR505000']
    ws.cell(row = 10, column = 2).value=calcData['CTR1005000']
    ws.cell(row = 1, column = 4).value=calcData['STC']

    #Graph

    chart2 = LineChart()
    xAxis = Reference(ws,min_col=3, min_row=27, max_col=3, max_row=47)
    # Rw
    values = Reference(ws,min_col=4, min_row=27, max_col=4, max_row=47)
    series = Series(values,title="R")
    chart2.append(series)
    # Rwref
    values = calcData['rwCurve']
    values = Reference(ws,min_col=7, min_row=2, max_col=27, max_row=2)
    series = Series(values,title="Curva de referencia")
    chart2.append(series)

    chart2.set_categories(xAxis)
    chart2.height = 10 # default is 7.5
    chart2.width = 15 # default is 15
    chart2.x_axis.title = 'f[Hz]'
    chart2.y_axis.title = 'Db'
    chart2.legend.position = 't'
    ws.add_chart(chart2, 'F25') 




    chart3 = LineChart()
    # STC
    values = Reference(ws,min_col=16, min_row=27, max_col=16, max_row=47)
    series = Series(values,title="R")
    chart3.append(series)
    # Rwref
    values = Reference(ws,min_col=7, min_row=3, max_col=27, max_row=3)
    series = Series(values,title="Curva de referencia")
    chart3.append(series)

    chart3.set_categories(xAxis)
    chart3.height = 10 # default is 7.5
    chart3.width = 15 # default is 15
    chart3.x_axis.title = 'f[Hz]'
    chart3.y_axis.title = 'Db'
    chart3.legend.position = 't'
    ws.add_chart(chart3, 'R25') 





    wb.active = 0








    name=export_save(export)
    return name

