calcData= {}

def get_variables(filtrado,length,height,thickness,material):
    from openpyxl import load_workbook
    import openpyxl
    #Vector filtrado

    if filtrado == 'octavas' :
        calcData['filterVector'] = [31.5,63,125,250,500,1000,2000,4000,8000,16000]
        calcData['db'] = 0.707
        calcData['octave'] = 1
    else:
        calcData['filterVector'] = [20,25,31.5,40,50,63,80,100,125,160,200,250,315,400,500,630,800,1000,1250,1600,2000,2500,3150,4000,5000,6300,8000,10000,12500,16000,20000]
        calcData['db'] = 0.236
        calcData['octave'] = 3       

    calcData['isoVector']=[50,63,80,100,125,160,200,250,315,400,500,630,800,1000,1250,1600,2000,2500,3150,4000,5000]
    #Material
    #Lectura de excel de materiales y obtencion de parametros del material seleccionado

    materials = load_workbook(filename = 'resources/sheets/materials.xlsx')
    sheet_obj = materials.active 

    for cell in sheet_obj['C']:
        if(cell.value is not None):
            if material == cell.value:
                materialR= cell.row
                materialColumn= cell.column
        

    calcData['material'] = material
    calcData['density'] = (sheet_obj.cell(row = materialR, column = (cell.column + 1)).value)
    calcData['young'] = (sheet_obj.cell(row = materialR, column = (cell.column + 2)).value)
    calcData['lossFactor'] = (sheet_obj.cell(row = materialR, column = (cell.column + 3)).value)
    calcData['poisson'] = (sheet_obj.cell(row = materialR, column = (cell.column + 4)).value)
    
 

    #Largo
    calcData['length'] = length
    #Altura
    calcData['height'] = height
    #Espesor
    calcData['thickness'] = thickness

    
  

    #Métodos


    return calcData
    